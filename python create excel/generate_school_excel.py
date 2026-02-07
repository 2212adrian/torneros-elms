from pathlib import Path
import random
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

STUDENTS_FILE = "Students_Report.xlsx"
GRADES_FILE = "Grades_Report.xlsx"
OUTPUT_FILE = "Unified_Student_System.xlsx"

# --- Helper Functions ---

def header_map(ws):
    headers = [cell.value for cell in ws[1]]
    return {h: i for i, h in enumerate(headers) if h is not None}

def get_value(row, idx, key, default=None):
    col = idx.get(key)
    if col is None:
        return default
    return row[col]

def table_ref(col_end, data_len):
    last_row = max(2, 1 + data_len)
    return f"A1:{col_end}{last_row}", last_row

def table_ref_at(start_row, col_end, data_len):
    last_row = max(start_row + 1, start_row + data_len)
    return f"A{start_row}:{col_end}{last_row}", last_row

def random_idcode(existing, digits=6):
    lower = 10 ** (digits - 1)
    upper = 10 ** digits - 1
    while True:
        value = str(random.randint(lower, upper))
        if value not in existing:
            existing.add(value)
            return value

def sample_names():
    first = [
        "Alex", "Bea", "Carlo", "Dina", "Eli", "Faye", "Gio", "Hana",
        "Ivan", "Jade", "Kira", "Liam", "Mia", "Nico", "Omar", "Pia",
        "Quin", "Rosa", "Sara", "Troy", "Uma", "Vera", "Wes", "Xena",
        "Yani", "Zane"
    ]
    last = [
        "Santos", "Garcia", "Reyes", "Cruz", "Lopez", "Flores", "Ramos",
        "Gonzales", "Torres", "Rivera", "Mendoza", "Aquino", "Morales",
        "Castro", "Delos Reyes", "Perez"
    ]
    names = []
    for f in first:
        for l in last:
            names.append(f"{f} {l}")
    return names

def fake_contact(existing):
    while True:
        value = f"09{random.randint(100000000, 999999999)}"
        if value not in existing:
            existing.add(value)
            return value

def fake_email(id_code):
    return f"student{id_code}@torneros.edu.ph"

def compute_age_formula(row):
    return f'=IF(D{row}="","",DATEDIF(D{row},TODAY(),"Y"))'

def normalize_year_level(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    mapping = {
        "1": "1st year", "2": "2nd year", "3": "3rd year", "4": "4th year",
        "1st": "1st year", "2nd": "2nd year", "3rd": "3rd year", "4th": "4th year",
        "first": "1st year", "second": "2nd year", "third": "3rd year", "fourth": "4th year",
    }
    key = text.lower()
    return mapping.get(key, text)

# --- File Loading ---
students_path = Path(STUDENTS_FILE)
grades_path = Path(GRADES_FILE)
if not students_path.exists() or not grades_path.exists():
    missing = []
    if not students_path.exists():
        missing.append(STUDENTS_FILE)
    if not grades_path.exists():
        missing.append(GRADES_FILE)
    raise FileNotFoundError(f"Missing input file(s): {', '.join(missing)}")

students_wb = load_workbook(students_path)
grades_wb = load_workbook(grades_path)

students_ws = students_wb.active
grades_ws = grades_wb.active

students_idx = header_map(students_ws)
grades_idx = header_map(grades_ws)

wb = Workbook()
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

# --- Students Sheet ---
ws = wb.active
ws.title = "Students"
ws.append([
    "student_id", "fullName", "gender", "birthdate",
    "age", "course", "contactNumber", "email"
])

student_lookup = {}
idcode_map = {}
existing_idcodes = set()
name_pool = sample_names()
name_cursor = 0
existing_contacts = set()
students_rows_processed = []

for row in students_ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue

    id_code_raw = get_value(row, students_idx, "idCode")
    id_code = str(id_code_raw) if id_code_raw else None
    if not id_code or id_code in existing_idcodes:
        id_code = random_idcode(existing_idcodes, digits=6)
    else:
        existing_idcodes.add(id_code)
    if id_code_raw:
        idcode_map[str(id_code_raw)] = id_code

    if name_cursor < len(name_pool):
        full_name = name_pool[name_cursor]
        name_cursor += 1
    else:
        full_name = f"Student {len(students_rows_processed) + 1}"

    gender_raw = get_value(row, students_idx, "gender")
    if isinstance(gender_raw, str):
        gender_key = gender_raw.strip().lower()
        gender = "Male" if gender_key == "male" else "Female" if gender_key == "female" else None
    elif isinstance(gender_raw, (int, float)):
        gender = "Male" if int(gender_raw) == 1 else "Female" if int(gender_raw) == 2 else None
    else:
        gender = None
    birthdate = get_value(row, students_idx, "birthdate")
    age = None
    section = get_value(row, students_idx, "section")
    course = None
    if section:
        course = str(section).split("-")[0].strip()
    contact_number = fake_contact(existing_contacts)
    email = fake_email(id_code)

    student_lookup[id_code] = full_name
    students_rows_processed.append([
        id_code, full_name, gender, birthdate,
        age, course, contact_number, email
    ])

for row in students_rows_processed:
    ws.append(row)

students_table = Table(displayName="StudentsTable", ref=table_ref("H", len(students_rows_processed))[0])
students_table.tableStyleInfo = style
ws.add_table(students_table)

for r in range(2, 2 + len(students_rows_processed)):
    ws[f"A{r}"].number_format = "@"
    ws[f"E{r}"] = compute_age_formula(r)

gender_dd = DataValidation(type="list", formula1='"Male,Female"', allow_blank=True)
ws.add_data_validation(gender_dd)
gender_dd.add("C2:C5000")

# --- Subjects pool (random per grade row) ---
subject_pool = [
    "Introduction to Computing",
    "Programming 1",
    "Discrete Mathematics",
    "Data Structures",
    "Database Systems",
    "Software Engineering",
    "Web Development",
    "Mobile App Dev",
    "Artificial Intelligence",
]

# --- Grades Sheet ---
ws4 = wb.create_sheet("Grades")
ws4.append([
    "student_id", "subject_name",
    "prelim", "midterm", "prefinal", "finals",
    "Final Average", "GPA", "Remarks",
    "schoolYear", "semester",
    "_student_id_key"
])

grades_rows_processed = []
for row in grades_ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    id_code = get_value(row, grades_idx, "studentIdCode")
    student_id = idcode_map.get(str(id_code)) if id_code else None
    if not student_id:
        continue
    subject_name = random.choice(subject_pool)

    prelim = get_value(row, grades_idx, "prelim")
    midterm = get_value(row, grades_idx, "midterm")
    prefinal = get_value(row, grades_idx, "prefinal")
    finals = get_value(row, grades_idx, "finals")
    school_year = get_value(row, grades_idx, "schoolYear")
    semester = get_value(row, grades_idx, "semester")

    grades_rows_processed.append([
        None, subject_name, prelim, midterm, prefinal, finals,
        None, None, None,
        school_year, semester,
        student_id
    ])

for row in grades_rows_processed:
    ws4.append(row)

last_grade_row_num = max(2, 1 + len(grades_rows_processed))

# Add formulas for Full Name, Subject Name, Average, GPA, Remarks
for r in range(2, last_grade_row_num + 1):
    ws4[f"A{r}"] = f'=IFERROR(VLOOKUP(L{r},StudentsTable,2,FALSE),"Student Not Found")'
    ws4[f"G{r}"] = f"=AVERAGE(C{r}:F{r})"
    ws4[f"H{r}"] = (
        f"=IF(G{r}>=96,1,"
        f"IF(G{r}>=94,1.25,"
        f"IF(G{r}>=91,1.5,"
        f"IF(G{r}>=89,1.75,"
        f"IF(G{r}>=86,2,"
        f"IF(G{r}>=83,2.25,"
        f"IF(G{r}>=80,2.5,"
        f"IF(G{r}>=77,2.75,"
        f"IF(G{r}>=75,3,"
        f"IF(G{r}>=70,4,5))))))))))"
    )
    ws4[f"I{r}"] = f'=IF(H{r}<=3,"PASSED","FAILED")'

grades_table = Table(displayName="GradesTable", ref=table_ref("I", len(grades_rows_processed))[0])
grades_table.tableStyleInfo = style
ws4.add_table(grades_table)

# --- Styling ---
id_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
auto_fill = PatternFill(fill_type="solid", fgColor="E6D6FF")
passed_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
failed_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
gender_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")

def color_columns(ws_target, col_letters, data_len, fill):
    for col in col_letters:
        for row in range(1, 2 + data_len):
            ws_target[f"{col}{row}"].fill = fill

color_columns(ws, ["A", "E"], len(students_rows_processed), id_fill)
color_columns(ws, ["C"], len(students_rows_processed), gender_fill)

ws4.column_dimensions["L"].hidden = True
for r in range(2, last_grade_row_num + 1):
    ws4[f"L{r}"].number_format = "@"

color_columns(ws4, ["G", "H"], len(grades_rows_processed), auto_fill)

passed_rule = FormulaRule(formula=[f'$I$2="PASSED"'], fill=passed_fill, stopIfTrue=True)
failed_rule = FormulaRule(formula=[f'$I$2="FAILED"'], fill=failed_fill, stopIfTrue=True)
ws4.conditional_formatting.add(f"I2:I{last_grade_row_num}", passed_rule)
ws4.conditional_formatting.add(f"I2:I{last_grade_row_num}", failed_rule)
ws4.conditional_formatting.add(f"J2:J{last_grade_row_num}", passed_rule)
ws4.conditional_formatting.add(f"J2:J{last_grade_row_num}", failed_rule)
ws4.conditional_formatting.add(f"K2:K{last_grade_row_num}", passed_rule)
ws4.conditional_formatting.add(f"K2:K{last_grade_row_num}", failed_rule)

# --- Save ---
wb.save(OUTPUT_FILE)
print(f"SUCCESS! Unified file created: {OUTPUT_FILE}")
